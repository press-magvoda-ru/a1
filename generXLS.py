#import pdb; pdb.set_trace()
# breakpoint()

import openpyxl
import os
import sys
import reparseWxMx
from reparseWxMx import PgIsEmpy,bundle,lstInWithExtention,Deliveries,defMekDeliv
#import fitz
import rezname
import string
from collections import Counter
alf=string.ascii_uppercase
typefilesOfdata=".dict_of_pages"
ZZ=[0]
"""Urs по мере увелечения контрАгентов выбирать из Gui c нижеследующим умолчанием"""
Urs='ВОДОКАНАЛ&ТЕПЛОФИКАЦИЯ&МЭК&АЗИМУТ&ГКС МКД'
NeedWF=0
OnlyTotals=1
def mainXLSsheetAndFresh(mnL, wb,wf): ## пока один поток сбора в два потока выгруза
    #фарш#
    """
    wf - выход для фреша - выкатки номеров пустых по файл-вкладка
    """
    #open dict of key-filename value=
    # all Pgs 
    # 
    # info about all (W|M)
    strTotal='ИТОГОВАЯ'; fl=0; 
    if wf:
        wf.create_sheet(strTotal);  total=wf[strTotal]
    
    str_pgTotal ='total';           wb.create_sheet(str_pgTotal);   pgTotal=wb[str_pgTotal]
    isFst=True; pgTotR=0
    strFrsh='oFrsh';                wb.create_sheet(strFrsh);       pgFrsh=wb[strFrsh]
    1;          pgFrshR=1
    Kontrs=Counter()
    for fullPathFile in mnL:
        if fullPathFile.find('$bundle')<0:
            continue
        bI=1;   bT=bI+2;    bC=1+1 # len bundle is 6
        r,c,freR,freC = 0,0,0,1,
        def put(c, Thing):            
            if not OnlyTotals:  sheet.cell(row=r, column=c).value = str(Thing)
        def putAndFrsh(c, Thing, u=True):
            put(c,Thing)
            if u:
                putPgToFrsh(c+1,Thing)

        def putfre(freR,Thing):
            if wf:
                shfre.cell(row=freR+bI, column=freC).value = str(Thing)
                putTot(freR+bT,fl,Thing)
        def putTot(r,c,Thing):
            if wf:
                total.cell(row=r, column=c).value = str(Thing)
        def putPgTot(c,Thing):          pgTotal.cell(row=pgTotR, column=c).value = str(Thing)
        
        def putPgToFrsh(c,Thing):       pgFrsh.cell(row=pgFrshR, column=c).value = str(Thing)


        def putAndPgTot(c,Thing, u=True):
            put(c,Thing)
            if u:
                putPgTot(c+1,Thing)

        w1, w2, wSqsW, wpaW ,wSqsM,wNums = [0]*6
        c1, cSqsW,cuuuW,cpaW, cpaM,cuuuM,cSqsM,c2 = [0]*8
        (pdfname :=(sp:=fullPathFile.split('$'))[0].split('\\')[-1].strip())
        stat=sp[1].replace(typefilesOfdata,'')
        print(f'{stat=}');  stat=eval(stat) #ОЛОЛО опаснАсте тся ться гауби
        
        OnlyTotals or  wb.create_sheet(pdfname);
        wf and wf.create_sheet(pdfname)
        
        pdfPath='$'.join(sp[:-1])
        print(fl:=fl+1,fullPathFile)#, end=' ') print(pdfPath)
        putTot(2,fl,pdfname)

        #bundle = namedtuple('bundle','m w P kvt sps tot')
        sheet = OnlyTotals or wb[pdfname]
        shfre = wf and wf[pdfname]
        def Hehehe(n): # 1-26 A-Z; 27-... AB ... for first 26**2 ?
            if (n:=n-1)<26:return alf[n]
            return alf[n//26-1]+alf[n%26]   # n явно меньше 625
        if wf:
            total.column_dimensions[Hehehe(fl)].width = 12
            
        agre=14; 

        r+=1
        for i in 'b':
            putAndFrsh(1,f'толькоМЭК',isFst)
            putAndFrsh(2,f'толькоВдТ',isFst) #
            putAndFrsh(3,f'двухСторо',isFst)
            putAndFrsh(4,f'стрСодерж',isFst)
            putAndFrsh(5,f'стрПустых',isFst)
            putAndFrsh(6,f'стрОбщее',isFst)
        r+=1
        pgFrshR+=1; putPgToFrsh(1,pdfname)    
        for i in 'c':
            putAndFrsh(1,f'{stat.m:^9}')
            putAndFrsh(2,f'{stat.w:^9}')
            putAndFrsh(3,f'{stat.P:^9}')
            putAndFrsh(4,f'{stat.kvt:^9}')
            putAndFrsh(5,f'{stat.sps:^9}')
            putAndFrsh(6,f'{stat.tot:^9}')
        
        #print(wb.sheetnames)
        #sheet.cell(row=(r := r+1), column=(c := 2)).value=f'{sp[-1]}:';wNums=len(sp[-1])+1
        
        Pgs=reparseWxMx.DictFromFile(fullPathFile)
        #os.system(f'del "{fullPathFile}"')
        #set header
        r+=1
        if isFst: pgTotR+=1
        for i in 'a':
            putAndPgTot(c:=c+1,"W-№", isFst);
            putAndPgTot(c:=c+1,"W-ЛС", isFst);
            putAndPgTot(c:=c+1,"W-Адрес", isFst);
            putAndPgTot(c:=c+1,"W-площадь", isFst);
            putAndPgTot(c:=c+1,"W-ФИО", isFst);
            putAndPgTot(c:=c+1,"W-ЕЛС", isFst);
    
            putAndPgTot(c:=c+1,"M-ЕЛС", isFst);
            putAndPgTot(c:=c+1,"M-ФИО", isFst);
            putAndPgTot(c:=c+1,"M-площадь", isFst);
            putAndPgTot(c:=c+1,"M-Адрес", isFst);
            putAndPgTot(c:=c+1,"M-ЛС", isFst);
            putAndPgTot(c:=c+1,"M-№", isFst);
            
            putAndPgTot(c:=c+1,"Доставщик", isFst)

            putAndPgTot(c:=c+1,"ВсеКонтрА", isFst); agre=c;
            for j in Urs.split('&'):
                putAndPgTot(c:=c+1,f'{j}:', isFst)
        isFst=False; # для однократного  заголовка в total
        sz=c    
        for x,y in Pgs:
            PgIsEmpy(x) and putfre(freR:=freR+1,x.pN+1)
            PgIsEmpy(y) and putfre(freR:=freR+1,y.pN+1)
            r+=1;  l = 0
            # cell = sheet.cell(row=(r := r+1), column=(c := c+1))
            # els = x.els
            # cell.value = f'=HYPERLINK("{pdfPath}.pdf","{els}")'
            # cell.style = "Hyperlink"
            pgTotR+=1; putPgTot(1,pdfname);
            putAndPgTot(l := l+1, f'={x.pN+1}')
            putAndPgTot(cpaW:=(l := l+1), x.pa); wpaW=max(wpaW,len(x.pa))
            putAndPgTot(c1 := (l := l+1), x.adr.replace('%', '/').rjust(w1 := max(w1, len(x.adr))))
            putAndPgTot(cSqsW := (l := l+1), x.sq)#.rjust
            (wSqsW := max(wSqsW, len(x.sq)))#)
            putAndPgTot(cuuuW:=(l := l+1), x.u)
            putAndPgTot(l := l+1, x.els)
            z=y
            putAndPgTot(l := l+1, z.els)
            putAndPgTot(cuuuM:=( l:= l+1), z.u)
            putAndPgTot(cSqsM:=(l := l+1), z.sq.rjust(wSqsM:=max(wSqsM,len(z.sq))))
            putAndPgTot(l := l+1, z.adr.replace('%', '/'))
            w2, c2 = max(w2, len(z.adr)), l
            putAndPgTot(cpaM := (l := l+1), z.pa)
            putAndPgTot(l := l+1, f'={z.pN+1}')
            
            (D:=x.Deliv or y.Deliv)
            for i in Deliveries.split('&'):
                if ~D.find(i):
                    DD=i;break;
            else:
                    DD=D#defMekDeliv
            putAndPgTot(l:=l+1,DD) # выбор из управляек и фолбэк Ливицкая здесь или при сборе ?
            Kontrs[V:='&'.join([x.UrFcs,y.UrFcs]).strip('&')]+=1
            putAndPgTot(l:=l+1,V)
            for j in Urs.split('&'):
                putAndPgTot(l:=l+1,f'={int(not not ~V.find(j))}')  #"пока так" если имена могут префиксоватся что искать &ndl& с предварительным оборачиванием &V& 
        if wf:
            putfre(0,f'{freR}');
            shfre.column_dimensions[alf[0]].width = 20 #Кол-во: в левленный столбец пояснение если чё
        ZZ[0]+=freR

        if not OnlyTotals:
            for i in range(sz):
                sheet.column_dimensions[alf[i]].width = 13
            for i in [2,9]:
                sheet.column_dimensions[alf[i]].width = 40
        continue
        sheet.column_dimensions['A'].width = 13
        sheet.column_dimensions[alf[1]].width = wNums+1 #len(str(w))+1
        sheet.column_dimensions[alf[c1-1]].width = w1
        sheet.column_dimensions[alf[cSqsW-1]].width = wSqsW
        sheet.column_dimensions[alf[cuuuW-1]].width = 3+3
        sheet.column_dimensions[alf[cpaW-1]].width = wpaW+1.1
        
        sheet.column_dimensions[alf[cpaM-1]].width = len(z.pa)*1.2
        sheet.column_dimensions[alf[cuuuM-1]].width = 3+3
        sheet.column_dimensions[alf[cSqsM-1]].width = wSqsM
        sheet.column_dimensions[alf[c2-1]].width = w2 
        sheet.column_dimensions[alf[l-1]].width = len(f'{100}')+1
    #выхлоп oBuhm:
    def toBuhm(): # за локальность имён!
        strBuhm='oBuhm';                wb.create_sheet(strBuhm);       pgBuhm=wb[strBuhm]
        for i,(k,v) in enumerate(Kontrs.items(),1):
            pgBuhm.cell(row=i, column=1).value = str(k)
            pgBuhm.cell(row=i, column=2).value = str(v)
    toBuhm()
    putTot(1,1,f'{ZZ[0]}') #Общее кол-во:  в левленный столбец пояснение если чё
def makeXLS(path):
    #s=f'*{typefilesOfdata}'
    (mnL := #sorted(
   #os.popen(f'dir "{os.path.join(path,s)}" /S /O-S /b').read().splitlines()
    lstInWithExtention(path,ext=typefilesOfdata)
    ) # /OD get Size of pdf desc
    """ """
    #)
    wb = openpyxl.Workbook(); wb.remove_sheet(wb.active)
    wf = NeedWF and openpyxl.Workbook();    wf and wf.remove_sheet(wf.active)
    mainXLSsheetAndFresh(mnL, wb,wf)
    tik=rezname.rezname()
    nm =f'SURV$${tik}.xlsx' ;#ЫГКМ  ну теперь ФСЁ ясНО # и выживальщики и pop-surv.gov74.ru lol
    nf =wf and f'Fresh${tik}.xlsx'
    wb.save(nm:=f'{os.path.join(path,nm)}');
    wf and wf.save(nf:=f'{os.path.join(path,nf)}')
    #os.system(f'start "" "cmd /c {nm}"')
    return nm, nf

if __name__=='__main__':
    print(sys.argv)
    cur = (args := sys.argv)[1:] and args[1] or "."
    makeXLS(cur)

